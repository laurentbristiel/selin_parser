from openpyxl import load_workbook
import sys
import string

class selin_parser:
    """
    INSTALL
    Those steps are customized for Windows - but tested only on Mac OS X
    - install Python 2.x (last version)
        https://www.python.org/downloads/release/python-278/
    - install easy_install with Python
        https://adesquared.wordpress.com/2013/07/07/setting-up-python-and-easy_install-on-windows-7/
    - install openpyxl library with easy_install
        in a DOS : easy_install openpyxl
    - put this file (selin_parser.py) and the XLXS modifiers file in the same directory
    - in a DOS shell window, launch:
        python selin_parser.py my_modifier_file.xlsx
        (note: ignore the warning about "Discarded range with reserved name")
        python selin_parser.py my_modifier_file.xlsx >output.txt
        => in the current directory, the file "output.txt" can be used to fill the mod txt file. 
    """

    def __init__(self):
        # row 4 contains the name of the modifiers
        self._modifiers_name_row = 4
        # this range contains the list of all the characters modifiers
        self.char_modifier_col_begin = 'CD'
        self.char_modifier_col_end = 'DR'
        # this range contains the list of all the other modifiers
        self.other_modifier_col_begin = 'L'
        self.other_modifier_col_end = 'CB'
        # column 4 contains the name/code of the religions
        self._religion_name_col = 'D'
        # this range contains the list of all the religion
        self._religion_row_begin = 8
        self._religion_row_end = 279

    def parse_modifiers_sheet(self, file_path):
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.get_sheet_by_name('Definition')

        modifiers_file = open ('modifiers.txt', 'w')

        for religion_row in range(self._religion_row_begin, self._religion_row_end):
            modifiers_file.write("#################################################\n")
            modifiers_file.write(ws.cell(column=self.col2num(self._religion_name_col), row=religion_row).value+'\n')
            modifiers_file.write("#################################################\n")
            self.write_modifiers(modifiers_file, ws, 'character_modifier', religion_row,
                                 self.col2num(self.char_modifier_col_begin),
                                 self.col2num(self.char_modifier_col_end))
            self.write_modifiers(modifiers_file, ws, 'other_modifier', religion_row,
                                 self.col2num(self.other_modifier_col_begin),
                                 self.col2num(self.other_modifier_col_end))

    def write_modifiers(self, modifiers_file, ws, range_name, religion_row, begin, end):
        line_prefix = '\t\t'
        modifiers_file.write(line_prefix + range_name + ' = {\n')
        for col in range(begin, end+1):
            header_value = ws.cell(column=col, row=self._modifiers_name_row).value
            cell_value = ws.cell(column=col, row=religion_row).value
            if header_value is None or cell_value == 0:
                continue
            if isinstance(cell_value, float):
                modifiers_file.write(line_prefix + '\t' + header_value.encode('utf-8') + " = %0.2f\n" % cell_value)
            if isinstance(cell_value, int):
                modifiers_file.write(line_prefix + '\t' + header_value.encode('utf-8') + " = " + str(cell_value) + '\n')
        modifiers_file.write(line_prefix + '}\n')

    def col2num(self, col):
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num    

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print "usage: python selin_parser.py filename.xlsx"
    else:
        selin_parser = selin_parser()
        selin_parser.parse_modifiers_sheet(sys.argv[1])